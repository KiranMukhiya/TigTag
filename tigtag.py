#!/usr/bin/env python3
"""
TigTag
======
A single-file desktop GUI for generating "Tags" / "Sub-tags" metadata
columns across one or more Excel workbooks and worksheets.

Features:
  * Keyword Detection panel (checkboxes, select-all/clear-all, custom
    entries, separate Tag / Sub-tag sections)
  * Importable/exportable keyword Blacklist and Whitelist
  * Male/Female auto-detection with an importable/exportable custom name list
  * File-type validation (only rows with a recognized extension are tagged)
  * Multi-sheet, multi-file batch processing
  * Output saved next to the source file as "<name>_processed.xlsx"

Run:
    python tigtag.py

Requirements:
    pip install openpyxl
    (tkinter ships with most desktop Python installs; on Linux you may need
     `sudo apt install python3-tk`)

Copyright (c) 2026 Kiran Mukhiya. All rights reserved.
"""

import os
import re
import json
import sys
import threading
import traceback
import webbrowser
from pathlib import Path
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
except ImportError:
    openpyxl = None


APP_NAME = "TigTag"
APP_VERSION = "v1.0"
APP_TITLE = f"{APP_NAME} {APP_VERSION}"
APP_AUTHOR = "Kiran Mukhiya"
APP_AUTHOR_EMAIL = "kiranmkya@gmail.com"
APP_COPYRIGHT = f"\u00A9 {datetime.now().year} {APP_AUTHOR}. All rights reserved."

# Directory to look for bundled assets (the icon) in. When frozen by
# PyInstaller, __file__ doesn't point at the real asset location, so use
# sys._MEIPASS (onefile) / sys.executable's folder (onedir) instead.
if getattr(sys, "frozen", False):
    APP_DIR = Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
else:
    APP_DIR = Path(__file__).resolve().parent
ICON_PATH = APP_DIR / "tigtag.png"

# =====================================================================
# THEME — primary olive green, accent deep forest green, light overall theme
# =====================================================================

PRIMARY = "#6b8e23"        # olive green — panels, tabs, secondary buttons
ACCENT = "#294a41"         # deep forest green — primary actions, emphasis
BG = "#f3f6f5"             # light app background
SURFACE = "#ffffff"        # cards / listboxes / text entry surfaces
TEXT = "#20302b"           # body text (dark, readable on light bg)
TEXT_ON_ACCENT = "#ffffff"
BORDER = "#d3d9c3"         # soft border derived from primary
PRIMARY_LIGHT = "#eaeedd"  # very light tint of primary, for troughs/hover


def apply_theme(root):
    """Configures a light theme using the primary/accent palette across all
    ttk widgets. Classic tk widgets (Listbox, Text, Canvas) are styled
    individually at creation time via the helper functions below."""
    root.configure(bg=BG)
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    style.configure(".", background=BG, foreground=TEXT)
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=TEXT)
    style.configure("TLabelframe", background=BG, foreground=ACCENT,
                     bordercolor=BORDER, relief="groove")
    style.configure("TLabelframe.Label", background=BG, foreground=ACCENT,
                     font=("TkDefaultFont", 10, "bold"))

    style.configure("TCheckbutton", background=BG, foreground=TEXT)
    style.map("TCheckbutton",
              background=[("active", BG)],
              foreground=[("disabled", "#9aa8a3")])

    style.configure("TButton", background=PRIMARY, foreground=TEXT_ON_ACCENT,
                     borderwidth=0, focuscolor=PRIMARY, padding=6)
    style.map("TButton",
              background=[("active", ACCENT), ("pressed", ACCENT), ("disabled", PRIMARY_LIGHT)],
              foreground=[("disabled", "#7c8a85")])

    style.configure("Accent.TButton", background=ACCENT, foreground=TEXT_ON_ACCENT,
                     borderwidth=0, focuscolor=ACCENT, padding=7)
    style.map("Accent.TButton",
              background=[("active", PRIMARY), ("pressed", PRIMARY)])

    # Top-bar buttons: inactive/hover colors swapped relative to the
    # standard TButton above (inactive = accent, hover = primary).
    style.configure("TopBar.TButton", background=ACCENT, foreground=TEXT_ON_ACCENT,
                     borderwidth=0, focuscolor=ACCENT, padding=6)
    style.map("TopBar.TButton",
              background=[("active", PRIMARY), ("pressed", PRIMARY), ("disabled", PRIMARY_LIGHT)],
              foreground=[("disabled", "#7c8a85")])

    style.configure("TNotebook", background=BG, bordercolor=BORDER)
    style.configure("TNotebook.Tab", background=PRIMARY_LIGHT, foreground=ACCENT,
                     padding=(14, 7), borderwidth=0)
    style.map("TNotebook.Tab",
              background=[("selected", ACCENT)],
              foreground=[("selected", TEXT_ON_ACCENT)])

    style.configure("TEntry", fieldbackground=SURFACE, foreground=TEXT,
                     bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER)
    style.configure("TCombobox", fieldbackground=SURFACE, foreground=TEXT,
                     background=PRIMARY, bordercolor=BORDER, arrowcolor=ACCENT)
    style.map("TCombobox", fieldbackground=[("readonly", SURFACE)],
              foreground=[("readonly", TEXT)])

    style.configure("TProgressbar", background=ACCENT, troughcolor=PRIMARY_LIGHT,
                     bordercolor=BORDER, lightcolor=ACCENT, darkcolor=ACCENT)

    style.configure("Vertical.TScrollbar", background=PRIMARY, troughcolor=BG,
                     arrowcolor=ACCENT, bordercolor=BG)
    style.configure("Horizontal.TScrollbar", background=PRIMARY, troughcolor=BG,
                     arrowcolor=ACCENT, bordercolor=BG)
    style.configure("TSeparator", background=BORDER)


def style_listbox(lb: tk.Listbox):
    lb.configure(bg=SURFACE, fg=TEXT, selectbackground=ACCENT, selectforeground=TEXT_ON_ACCENT,
                 highlightbackground=BORDER, highlightcolor=ACCENT, relief="solid", bd=1)
    return lb


def style_text(widget: tk.Text):
    widget.configure(bg=SURFACE, fg=TEXT, insertbackground=ACCENT,
                      highlightbackground=BORDER, highlightcolor=ACCENT, relief="solid", bd=1)
    return widget


def style_canvas(canvas: tk.Canvas):
    canvas.configure(bg=BG, highlightthickness=0)
    return canvas


# =====================================================================
# CONFIG / PERSISTENCE
# =====================================================================

CONFIG_DIR = Path(os.path.expanduser("~")) / ".tigtag"
CONFIG_DIR.mkdir(parents=True, exist_ok=True)
FEEDBACK_LOG = CONFIG_DIR / "feedback.log"

DEFAULT_MALE_NAMES = ["Priyangshu", "Sandesh", "Shushant", "Krushil", "Sagar", "Krishna",
                       "Amaan", "Dhruv", "Sahil", "Subhash", "Noel"]
DEFAULT_FEMALE_NAMES = ["Liz", "Kiran", "Joe", "Mississippi", "Simran", "Anjitha", "Harini",
                         "Tanishka", "Shatakshi", "Manasi", "Diksha", "Anushka", "Kousar"]

DEFAULT_EXTENSIONS = [".txt", ".pdf", ".csv", ".pcm", ".wav", ".mp3", ".m4a", ".json",
                       ".xml", ".docx", ".xlsx", ".png", ".jpg", ".jpeg", ".flac", ".aac"]

# Tokens that should never be offered as tag/subtag candidates on their own.
GENERIC_STOPWORDS = {
    "delivery", "consent", "forms", "form", "done", "poc", "updated", "data", "year",
    "recordings", "recording", "annotation", "folder", "files", "file", "the", "and",
    "or", "of", "for", "with", "in", "on", "to", "a", "an", "new", "old", "final",
    "v1", "v2", "copy", "backup", "misc", "general", "common", "test", "temp", "draft",
}

CONFIG_FILES = {
    "blacklist": CONFIG_DIR / "blacklist.json",
    "whitelist": CONFIG_DIR / "whitelist.json",
    "male_names": CONFIG_DIR / "male_names.json",
    "female_names": CONFIG_DIR / "female_names.json",
    "extensions": CONFIG_DIR / "extensions.json",
    "keyword_rules": CONFIG_DIR / "keyword_rules.json",
}


def _load_json(path, default):
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default
    return default


def _save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


class ConfigStore:
    """Loads / saves all persistent lists to the local config directory
    (~/.tigtag). Every list here is importable / exportable from the GUI."""

    def __init__(self):
        self.blacklist = _load_json(CONFIG_FILES["blacklist"], {"tags": [], "subtags": []})
        self.whitelist = _load_json(CONFIG_FILES["whitelist"], {"tags": [], "subtags": []})
        self.male_names = _load_json(CONFIG_FILES["male_names"], list(DEFAULT_MALE_NAMES))
        self.female_names = _load_json(CONFIG_FILES["female_names"], list(DEFAULT_FEMALE_NAMES))
        self.extensions = _load_json(CONFIG_FILES["extensions"], list(DEFAULT_EXTENSIONS))
        # keyword_rules: { "<lowercase keyword>": {"display": str, "category": "tag"/"subtag",
        #                                           "enabled": bool, "outputs": [str, ...]} }
        self.keyword_rules = _load_json(CONFIG_FILES["keyword_rules"], {})

    def save_blacklist(self):
        _save_json(CONFIG_FILES["blacklist"], self.blacklist)

    def save_whitelist(self):
        _save_json(CONFIG_FILES["whitelist"], self.whitelist)

    def save_male_names(self):
        _save_json(CONFIG_FILES["male_names"], self.male_names)

    def save_female_names(self):
        _save_json(CONFIG_FILES["female_names"], self.female_names)

    def save_extensions(self):
        _save_json(CONFIG_FILES["extensions"], self.extensions)

    def save_keyword_rules(self):
        _save_json(CONFIG_FILES["keyword_rules"], self.keyword_rules)


# =====================================================================
# PATH TOKENIZING / KEYWORD DETECTION
# =====================================================================

DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),
]
IP_PATTERN = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
GENDER_TOKEN_PATTERN = re.compile(r"^(Male|Female)(\d+)?$", re.IGNORECASE)
GENDER_TOKEN_SEARCH = re.compile(r"(Male|Female)(\d+)?", re.IGNORECASE)


def split_camel_case(word):
    """MessageApp -> Message App (splits on lower->UPPER boundary only)."""
    out = []
    for i, ch in enumerate(word):
        if i > 0 and ch.isupper() and word[i - 1].islower():
            out.append(" ")
        out.append(ch)
    return "".join(out)


def is_date_like(token):
    return any(p.match(token) for p in DATE_PATTERNS)


def clean_token(token):
    token = split_camel_case(token)
    token = token.replace("-", " ").strip()
    return token


class PathAnalyzer:
    """Breaks a file / consent path into an ordered list of unique
    candidate keyword tokens, filtering generic noise, numbers, dates,
    IP-like share names, and known person names / gender words (those are
    handled separately by GenderDetector)."""

    def __init__(self, config: ConfigStore):
        self.config = config

    def _known_names(self):
        return {n.strip().lower() for n in (self.config.male_names + self.config.female_names) if n.strip()}

    def tokenize(self, path_str):
        if not path_str:
            return []
        raw = str(path_str).replace("\\", "/")
        segments = [s for s in raw.split("/") if s]
        known_names = self._known_names()
        tokens = []
        for seg in segments:
            if IP_PATTERN.match(seg):
                continue
            name_no_ext, ext = os.path.splitext(seg)
            base = name_no_ext if ext else seg
            for part in re.split(r"[_\s]+", base):
                part = part.strip(" .")
                if not part:
                    continue
                if part.isdigit():
                    continue
                if is_date_like(part):
                    continue
                low = part.lower()
                if low in GENERIC_STOPWORDS:
                    continue
                if low in known_names:
                    continue
                if GENDER_TOKEN_PATTERN.match(part):
                    continue
                cleaned = clean_token(part)
                if cleaned and cleaned.lower() not in GENERIC_STOPWORDS and len(cleaned) > 1:
                    tokens.append(cleaned)
        seen = set()
        ordered = []
        for t in tokens:
            k = t.lower()
            if k not in seen:
                seen.add(k)
                ordered.append(t)
        return ordered


def scan_candidates(config: ConfigStore, path_pairs):
    """path_pairs: iterable of (audio_path, consent_path).
    Returns {lowercase_keyword: {"display": str, "count": int}} across all rows."""
    analyzer = PathAnalyzer(config)
    counts = {}
    for audio_path, consent_path in path_pairs:
        toks = analyzer.tokenize(audio_path) + analyzer.tokenize(consent_path or "")
        local_seen = set()
        for tok in toks:
            key = tok.lower()
            if key in local_seen:
                continue
            local_seen.add(key)
            if key not in counts:
                counts[key] = {"display": tok, "count": 0}
            counts[key]["count"] += 1
    return counts


# =====================================================================
# GENDER DETECTION
# =====================================================================

class GenderDetector:
    """Auto-detects Male / Female from a row's paths. Priority:
    1) literal gender token in the path (e.g. 'Female13', 'Male_05')
    2) a known name (from the custom male/female name lists) appearing
       anywhere in the path (e.g. consent form file name)
    Returns '' if nothing is found (male/female tag is simply skipped)."""

    def __init__(self, config: ConfigStore):
        self.config = config

    def detect(self, *path_strings):
        combined = " / ".join(str(p) for p in path_strings if p)
        m = GENDER_TOKEN_SEARCH.search(combined)
        if m:
            return m.group(1).capitalize()
        low = combined.lower()
        for name in self.config.male_names:
            n = name.strip().lower()
            if n and n in low:
                return "Male"
        for name in self.config.female_names:
            n = name.strip().lower()
            if n and n in low:
                return "Female"
        return ""


# =====================================================================
# TAGGING ENGINE
# =====================================================================

class TaggingEngine:
    def __init__(self, config: ConfigStore):
        self.config = config
        self.analyzer = PathAnalyzer(config)
        self.gender = GenderDetector(config)

    def is_valid_file(self, audio_path):
        """True if audio_path has a recognized, configured file extension."""
        if not audio_path:
            return False
        ext = os.path.splitext(str(audio_path))[1].lower()
        if not ext:
            return False
        return ext in {e.lower() for e in self.config.extensions}

    def build_tags(self, audio_path, consent_path):
        """Returns (tags_str, subtags_str) for one row."""
        tags, subtags = ["Audio"], []

        candidates = self.analyzer.tokenize(audio_path) + self.analyzer.tokenize(consent_path or "")
        seen, ordered_candidates = set(), []
        for c in candidates:
            k = c.lower()
            if k not in seen:
                seen.add(k)
                ordered_candidates.append(c)

        for cand in ordered_candidates:
            rule = self.config.keyword_rules.get(cand.lower())
            if not rule or not rule.get("enabled"):
                continue
            outputs = rule.get("outputs") or [rule.get("display", cand)]
            target = tags if rule.get("category") == "tag" else subtags
            for o in outputs:
                o = o.strip()
                if o and o not in target:
                    target.append(o)

        combined_low = ((str(audio_path) if audio_path else "") + " " +
                         (str(consent_path) if consent_path else "")).lower()

        for w in self.config.whitelist.get("tags", []):
            if w and w.strip().lower() in combined_low and w not in tags:
                tags.append(w)
        for w in self.config.whitelist.get("subtags", []):
            if w and w.strip().lower() in combined_low and w not in subtags:
                subtags.append(w)

        g = self.gender.detect(audio_path, consent_path)
        if g and g not in subtags:
            subtags.append(g)

        bl_tags = {b.strip().lower() for b in self.config.blacklist.get("tags", [])}
        bl_subtags = {b.strip().lower() for b in self.config.blacklist.get("subtags", [])}
        tags = [t for t in tags if t.lower() not in bl_tags]
        subtags = [s for s in subtags if s.lower() not in bl_subtags]

        return ", ".join(tags), ", ".join(subtags)


# =====================================================================
# EXCEL READ / WRITE HELPERS
# =====================================================================

class ExcelProcessor:
    COL_AUDIO = 1
    COL_CONSENT = 2
    COL_TAGS = 3
    COL_SUBTAGS = 4

    @staticmethod
    def load_workbook(path):
        return openpyxl.load_workbook(path)

    @staticmethod
    def iter_rows(ws):
        """Yield (row_idx, audio_path, consent_path) for data rows (row 2+)."""
        last_row = ws.max_row
        for r in range(2, last_row + 1):
            a = ws.cell(row=r, column=ExcelProcessor.COL_AUDIO).value
            c = ws.cell(row=r, column=ExcelProcessor.COL_CONSENT).value
            a = str(a).strip() if a is not None else ""
            c = str(c).strip() if c is not None else ""
            if not a:
                continue
            yield r, a, c

    @staticmethod
    def ensure_headers(ws):
        for text, col in (("Tags", ExcelProcessor.COL_TAGS), ("Sub-tags", ExcelProcessor.COL_SUBTAGS)):
            cell = ws.cell(row=1, column=col)
            if not cell.value:
                cell.value = text

    @staticmethod
    def output_path_for(source_path):
        p = Path(source_path)
        return p.with_name(f"{p.stem}_processed{p.suffix}")


def process_workbook(file_path, sheet_names, config: ConfigStore, include_folder_rows=False, log=lambda s: None):
    """Processes the given sheets of one workbook in place and saves it as
    <name>_processed<ext> alongside the source file. Returns (out_path, processed_count, skipped_count)."""
    engine = TaggingEngine(config)
    wb = ExcelProcessor.load_workbook(file_path)
    processed, skipped = 0, 0
    targets = sheet_names or wb.sheetnames
    for sn in targets:
        if sn not in wb.sheetnames:
            log(f"  (sheet '{sn}' not found, skipped)")
            continue
        ws = wb[sn]
        ExcelProcessor.ensure_headers(ws)
        sheet_processed = 0
        for r, audio, consent in ExcelProcessor.iter_rows(ws):
            is_file = engine.is_valid_file(audio)
            if not is_file and not include_folder_rows:
                skipped += 1
                continue
            tags, subtags = engine.build_tags(audio, consent)
            ws.cell(row=r, column=ExcelProcessor.COL_TAGS).value = tags
            ws.cell(row=r, column=ExcelProcessor.COL_SUBTAGS).value = subtags
            processed += 1
            sheet_processed += 1
        log(f"  Sheet '{sn}': {sheet_processed} rows tagged")
    out_path = ExcelProcessor.output_path_for(file_path)
    wb.save(out_path)
    return out_path, processed, skipped


# =====================================================================
# Small reusable scrollable frame
# =====================================================================

class ScrollableFrame(ttk.Frame):
    def __init__(self, parent, height=320, **kwargs):
        super().__init__(parent, **kwargs)
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0, height=height)
        style_canvas(self.canvas)
        self.vbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)

        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.vbar.pack(side="right", fill="y")

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel, add="+")

    def _on_mousewheel(self, event):
        try:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass


# =====================================================================
# Detection Panel: keyword candidates -> Tag / Subtag
# =====================================================================

class DetectionPanel(tk.Toplevel):
    def __init__(self, master, config: ConfigStore, candidates):
        """candidates: {lower_keyword: {"display": str, "count": int}}"""
        super().__init__(master)
        self.configure(bg=BG)
        self.title("Keyword Detection")
        self.geometry("780x620")
        self.config_store = config
        self.candidates = candidates
        self.vars = {}  # lower_keyword -> {"enabled": BooleanVar, "category": StringVar, "display": StringVar}

        self._seed_rules()
        self._build_ui()

    def _seed_rules(self):
        """Make sure every candidate has a rule entry (defaults: subtag, disabled)."""
        for key, info in self.candidates.items():
            if key not in self.config_store.keyword_rules:
                self.config_store.keyword_rules[key] = {
                    "display": info["display"],
                    "category": "subtag",
                    "enabled": False,
                    "outputs": [info["display"]],
                }

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=(10, 0))
        ttk.Label(top, text="Non-generic keywords detected across the selected data. "
                             "Check the ones you want written out, choose Tag or Sub-tag, "
                             "and optionally edit the exact text that gets written.",
                  wraplength=740, justify="left").pack(anchor="w")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tag_frame = ScrollableFrame(self.notebook)
        self.subtag_frame = ScrollableFrame(self.notebook)
        self.notebook.add(self.tag_frame, text="Tags")
        self.notebook.add(self.subtag_frame, text="Sub-tags")

        self._render_rows()

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=(0, 5))
        ttk.Button(btns, text="Select All (this tab)", command=lambda: self._set_all(True)).pack(side="left")
        ttk.Button(btns, text="Clear All (this tab)", command=lambda: self._set_all(False)).pack(side="left", padx=6)

        custom = ttk.LabelFrame(self, text="Add custom keyword")
        custom.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Label(custom, text="Keyword:").grid(row=0, column=0, padx=5, pady=6, sticky="e")
        self.custom_kw = tk.StringVar()
        ttk.Entry(custom, textvariable=self.custom_kw, width=22).grid(row=0, column=1, padx=5)
        ttk.Label(custom, text="Category:").grid(row=0, column=2, padx=5, sticky="e")
        self.custom_cat = tk.StringVar(value="tag")
        ttk.Combobox(custom, textvariable=self.custom_cat, values=["tag", "subtag"],
                     width=10, state="readonly").grid(row=0, column=3, padx=5)
        ttk.Label(custom, text="Writes as (optional, comma-sep):").grid(row=0, column=4, padx=5, sticky="e")
        self.custom_out = tk.StringVar()
        ttk.Entry(custom, textvariable=self.custom_out, width=20).grid(row=0, column=5, padx=5)
        ttk.Button(custom, text="Add", command=self._add_custom).grid(row=0, column=6, padx=8)

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(bottom, text="Save & Close", style="Accent.TButton", command=self._save_close).pack(side="right")
        ttk.Button(bottom, text="Cancel", command=self.destroy).pack(side="right", padx=8)

    def _render_rows(self):
        for child in self.tag_frame.inner.winfo_children():
            child.destroy()
        for child in self.subtag_frame.inner.winfo_children():
            child.destroy()

        items = sorted(self.config_store.keyword_rules.items(),
                        key=lambda kv: -self.candidates.get(kv[0], {}).get("count", 0))

        for key, rule in items:
            info = self.candidates.get(key, {"display": rule.get("display", key), "count": 0})
            parent = self.tag_frame.inner if rule.get("category") == "tag" else self.subtag_frame.inner
            row = ttk.Frame(parent)
            row.pack(fill="x", anchor="w", pady=1, padx=4)

            enabled_var = tk.BooleanVar(value=rule.get("enabled", False))
            display_var = tk.StringVar(value=rule.get("display", info["display"]))
            category_var = tk.StringVar(value=rule.get("category", "subtag"))
            self.vars[key] = {"enabled": enabled_var, "display": display_var, "category": category_var}

            cb = ttk.Checkbutton(row, variable=enabled_var,
                                  command=lambda k=key: self._sync_rule(k))
            cb.pack(side="left")
            lbl = ttk.Label(row, text=f"({info.get('count', 0)}x)", width=7)
            lbl.pack(side="left")
            entry = ttk.Entry(row, textvariable=display_var, width=26)
            entry.pack(side="left", padx=4)
            entry.bind("<FocusOut>", lambda e, k=key: self._sync_rule(k))
            combo = ttk.Combobox(row, textvariable=category_var, values=["tag", "subtag"],
                                  width=8, state="readonly")
            combo.pack(side="left", padx=4)
            combo.bind("<<ComboboxSelected>>", lambda e, k=key: self._move(k))

    def _sync_rule(self, key):
        v = self.vars[key]
        rule = self.config_store.keyword_rules[key]
        rule["enabled"] = v["enabled"].get()
        new_display = v["display"].get().strip() or rule.get("display", key)
        rule["display"] = new_display
        rule["outputs"] = [x.strip() for x in new_display.split(",") if x.strip()]

    def _move(self, key):
        self._sync_rule(key)
        new_category = self.vars[key]["category"].get()
        self.config_store.keyword_rules[key]["category"] = new_category
        # Defer the rebuild until after this event callback finishes (we're
        # currently inside a callback fired by a widget that _render_rows is
        # about to destroy), then jump to the tab the keyword moved into so
        # it's clearly still there rather than looking like it vanished.
        target_tab = self.tag_frame if new_category == "tag" else self.subtag_frame
        self.after(1, lambda: self._render_rows_and_focus(target_tab))

    def _render_rows_and_focus(self, target_tab):
        self._render_rows()
        self.notebook.select(target_tab)

    def _set_all(self, value):
        current_tab = self.notebook.tab(self.notebook.select(), "text")
        cat = "tag" if current_tab == "Tags" else "subtag"
        for key, rule in self.config_store.keyword_rules.items():
            if rule.get("category") == cat:
                rule["enabled"] = value
        self._render_rows()

    def _add_custom(self):
        kw = self.custom_kw.get().strip()
        if not kw:
            messagebox.showwarning(APP_TITLE, "Enter a keyword first.")
            return
        key = kw.lower()
        outputs = [x.strip() for x in self.custom_out.get().split(",") if x.strip()] or [kw]
        self.config_store.keyword_rules[key] = {
            "display": kw,
            "category": self.custom_cat.get(),
            "enabled": True,
            "outputs": outputs,
        }
        if key not in self.candidates:
            self.candidates[key] = {"display": kw, "count": 0}
        self.custom_kw.set("")
        self.custom_out.set("")
        self._render_rows()

    def _save_close(self):
        for key in list(self.vars.keys()):
            self._sync_rule(key)
        self.config_store.save_keyword_rules()
        self.destroy()


# =====================================================================
# Generic two-column (Tags / Sub-tags) import/export list editor
# used for both Blacklist and Whitelist
# =====================================================================

class TagSubtagListDialog(tk.Toplevel):
    def __init__(self, master, config: ConfigStore, list_name):
        """list_name: 'blacklist' or 'whitelist'"""
        super().__init__(master)
        self.configure(bg=BG)
        self.config_store = config
        self.list_name = list_name
        self.data = getattr(config, list_name)
        self.title(f"{list_name.capitalize()} Manager")
        self.geometry("560x460")
        self._build_ui()

    def _build_ui(self):
        ttk.Label(self, text=f"Keywords here are {'excluded from' if self.list_name=='blacklist' else 'always force-included in'} "
                              f"the Tags / Sub-tags output.", wraplength=520).pack(anchor="w", padx=10, pady=(10, 4))

        cols = ttk.Frame(self)
        cols.pack(fill="both", expand=True, padx=10)

        self.tag_list = self._build_column(cols, "Tags", "tags")
        self.subtag_list = self._build_column(cols, "Sub-tags", "subtags")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Import (.json)", command=self._import).pack(side="left")
        ttk.Button(btns, text="Export (.json)", command=self._export).pack(side="left", padx=6)
        ttk.Button(btns, text="Save & Close", style="Accent.TButton", command=self._save_close).pack(side="right")
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side="right", padx=6)

    def _build_column(self, parent, title, key):
        frame = ttk.LabelFrame(parent, text=title)
        frame.pack(side="left", fill="both", expand=True, padx=5)

        lb = tk.Listbox(frame, height=14, selectmode="extended")
        style_listbox(lb)
        lb.pack(fill="both", expand=True, padx=5, pady=5)
        for item in self.data.get(key, []):
            lb.insert("end", item)

        entry_row = ttk.Frame(frame)
        entry_row.pack(fill="x", padx=5, pady=(0, 5))
        var = tk.StringVar()
        ttk.Entry(entry_row, textvariable=var).pack(side="left", fill="x", expand=True)
        ttk.Button(entry_row, text="Add", command=lambda: self._add(lb, var)).pack(side="left", padx=4)
        ttk.Button(entry_row, text="Remove selected",
                   command=lambda: self._remove(lb)).pack(side="left")
        return lb

    def _add(self, listbox, var):
        v = var.get().strip()
        if v:
            listbox.insert("end", v)
            var.set("")

    def _remove(self, listbox):
        for idx in reversed(listbox.curselection()):
            listbox.delete(idx)

    def _collect(self):
        return {
            "tags": list(self.tag_list.get(0, "end")),
            "subtags": list(self.subtag_list.get(0, "end")),
        }

    def _import(self):
        path = filedialog.askopenfilename(title="Import list", filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                incoming = json.load(f)
            for item in incoming.get("tags", []):
                self.tag_list.insert("end", item)
            for item in incoming.get("subtags", []):
                self.subtag_list.insert("end", item)
            messagebox.showinfo(APP_TITLE, "Import complete.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not import: {e}")

    def _export(self):
        path = filedialog.asksaveasfilename(title="Export list", defaultextension=".json",
                                             filetypes=[("JSON", "*.json")],
                                             initialfile=f"{self.list_name}.json")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self._collect(), f, indent=2, ensure_ascii=False)
            messagebox.showinfo(APP_TITLE, "Exported.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not export: {e}")

    def _save_close(self):
        setattr(self.config_store, self.list_name, self._collect())
        if self.list_name == "blacklist":
            self.config_store.save_blacklist()
        else:
            self.config_store.save_whitelist()
        self.destroy()


# =====================================================================
# Gender name-list dialog
# =====================================================================

class GenderNamesDialog(tk.Toplevel):
    def __init__(self, master, config: ConfigStore):
        super().__init__(master)
        self.configure(bg=BG)
        self.config_store = config
        self.title("Male / Female Name Lists")
        self.geometry("560x460")
        self._build_ui()

    def _build_ui(self):
        ttk.Label(self, text="Used as a fallback when a row has no literal 'Male'/'Female' token in its path "
                              "(e.g. matched from the consent-form file name). Auto-detection first looks for a "
                              "direct gender token; if none is found the row is skipped for that tag.",
                  wraplength=520, justify="left").pack(anchor="w", padx=10, pady=(10, 4))

        cols = ttk.Frame(self)
        cols.pack(fill="both", expand=True, padx=10)

        self.male_lb = self._build_column(cols, "Male names", self.config_store.male_names)
        self.female_lb = self._build_column(cols, "Female names", self.config_store.female_names)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Import (.json)", command=self._import).pack(side="left")
        ttk.Button(btns, text="Export (.json)", command=self._export).pack(side="left", padx=6)
        ttk.Button(btns, text="Save & Close", style="Accent.TButton", command=self._save_close).pack(side="right")
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side="right", padx=6)

    def _build_column(self, parent, title, values):
        frame = ttk.LabelFrame(parent, text=title)
        frame.pack(side="left", fill="both", expand=True, padx=5)
        lb = tk.Listbox(frame, height=14, selectmode="extended")
        style_listbox(lb)
        lb.pack(fill="both", expand=True, padx=5, pady=5)
        for v in values:
            lb.insert("end", v)
        entry_row = ttk.Frame(frame)
        entry_row.pack(fill="x", padx=5, pady=(0, 5))
        var = tk.StringVar()
        ttk.Entry(entry_row, textvariable=var).pack(side="left", fill="x", expand=True)
        ttk.Button(entry_row, text="Add", command=lambda: self._add(lb, var)).pack(side="left", padx=4)
        ttk.Button(entry_row, text="Remove", command=lambda: self._remove(lb)).pack(side="left")
        return lb

    def _add(self, lb, var):
        v = var.get().strip()
        if v:
            lb.insert("end", v)
            var.set("")

    def _remove(self, lb):
        for idx in reversed(lb.curselection()):
            lb.delete(idx)

    def _import(self):
        path = filedialog.askopenfilename(title="Import names", filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                incoming = json.load(f)
            for v in incoming.get("male", []):
                self.male_lb.insert("end", v)
            for v in incoming.get("female", []):
                self.female_lb.insert("end", v)
            messagebox.showinfo(APP_TITLE, "Import complete.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not import: {e}")

    def _export(self):
        path = filedialog.asksaveasfilename(title="Export names", defaultextension=".json",
                                             filetypes=[("JSON", "*.json")], initialfile="gender_names.json")
        if not path:
            return
        try:
            data = {"male": list(self.male_lb.get(0, "end")), "female": list(self.female_lb.get(0, "end"))}
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            messagebox.showinfo(APP_TITLE, "Exported.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not export: {e}")

    def _save_close(self):
        self.config_store.male_names = list(self.male_lb.get(0, "end"))
        self.config_store.female_names = list(self.female_lb.get(0, "end"))
        self.config_store.save_male_names()
        self.config_store.save_female_names()
        self.destroy()


# =====================================================================
# File extensions dialog
# =====================================================================

class ExtensionsDialog(tk.Toplevel):
    def __init__(self, master, config: ConfigStore):
        super().__init__(master)
        self.configure(bg=BG)
        self.config_store = config
        self.title("Valid File Types")
        self.geometry("380x420")
        self._build_ui()

    def _build_ui(self):
        ttk.Label(self, text="Only file paths ending in one of these extensions are treated as "
                              "taggable files. Rows without a matching extension (e.g. bare folder "
                              "paths) can optionally still get basic tags via the main window checkbox.",
                  wraplength=340, justify="left").pack(anchor="w", padx=10, pady=(10, 6))
        self.lb = tk.Listbox(self, height=14, selectmode="extended")
        style_listbox(self.lb)
        self.lb.pack(fill="both", expand=True, padx=10)
        for e in self.config_store.extensions:
            self.lb.insert("end", e)

        row = ttk.Frame(self)
        row.pack(fill="x", padx=10, pady=6)
        self.var = tk.StringVar()
        ttk.Entry(row, textvariable=self.var).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Add", command=self._add).pack(side="left", padx=4)
        ttk.Button(row, text="Remove", command=self._remove).pack(side="left")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Import (.json)", command=self._import).pack(side="left")
        ttk.Button(btns, text="Export (.json)", command=self._export).pack(side="left", padx=6)
        ttk.Button(btns, text="Save & Close", style="Accent.TButton", command=self._save_close).pack(side="right")
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side="right", padx=6)

    def _add(self):
        v = self.var.get().strip()
        if v:
            if not v.startswith("."):
                v = "." + v
            self.lb.insert("end", v.lower())
            self.var.set("")

    def _remove(self):
        for idx in reversed(self.lb.curselection()):
            self.lb.delete(idx)

    def _import(self):
        path = filedialog.askopenfilename(title="Import extensions", filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                incoming = json.load(f)
            for v in incoming:
                self.lb.insert("end", v)
            messagebox.showinfo(APP_TITLE, "Import complete.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not import: {e}")

    def _export(self):
        path = filedialog.asksaveasfilename(title="Export extensions", defaultextension=".json",
                                             filetypes=[("JSON", "*.json")], initialfile="extensions.json")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(list(self.lb.get(0, "end")), f, indent=2)
            messagebox.showinfo(APP_TITLE, "Exported.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not export: {e}")

    def _save_close(self):
        self.config_store.extensions = list(self.lb.get(0, "end"))
        self.config_store.save_extensions()
        self.destroy()


# =====================================================================
# About dialog
# =====================================================================

class AboutDialog(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.configure(bg=BG)
        self.title(f"About {APP_NAME}")
        self.resizable(False, False)
        self.geometry("380x340")

        self._icon_img = None
        if ICON_PATH.exists():
            try:
                img = tk.PhotoImage(file=str(ICON_PATH))
                # Downscale large source icons for a reasonable on-screen size.
                factor = max(1, img.width() // 96)
                if factor > 1:
                    img = img.subsample(factor, factor)
                self._icon_img = img
                tk.Label(self, image=self._icon_img, bg=BG).pack(pady=(20, 8))
            except Exception:
                pass

        tk.Label(self, text=APP_NAME, bg=BG, fg=ACCENT,
                 font=("TkDefaultFont", 16, "bold")).pack()
        tk.Label(self, text=APP_VERSION, bg=BG, fg=TEXT,
                 font=("TkDefaultFont", 10)).pack(pady=(0, 10))
        tk.Label(self, text="Generates Tags / Sub-tags metadata across\n"
                             "Excel workbooks from configurable keyword,\n"
                             "gender, and file-type rules.",
                 bg=BG, fg=TEXT, justify="center").pack(pady=(0, 14))
        tk.Label(self, text=APP_COPYRIGHT, bg=BG, fg=TEXT,
                 font=("TkDefaultFont", 9)).pack(pady=(0, 4))

        sanction_row = tk.Frame(self, bg=BG)
        sanction_row.pack(pady=(0, 14))
        tk.Label(sanction_row, text="All rights sanctioned to ", bg=BG, fg=TEXT,
                 font=("TkDefaultFont", 9)).pack(side="left")
        link = tk.Label(sanction_row, text=APP_AUTHOR, bg=BG, fg=ACCENT,
                         font=("TkDefaultFont", 9, "underline"), cursor="hand2")
        link.pack(side="left")
        tk.Label(sanction_row, text=".", bg=BG, fg=TEXT,
                 font=("TkDefaultFont", 9)).pack(side="left")
        link.bind("<Button-1>", lambda e: webbrowser.open(f"mailto:{APP_AUTHOR_EMAIL}"))
        link.bind("<Enter>", lambda e: link.configure(fg=PRIMARY))
        link.bind("<Leave>", lambda e: link.configure(fg=ACCENT))
        ttk.Button(self, text="Close", style="Accent.TButton", command=self.destroy).pack(pady=(0, 16))


# =====================================================================
# Feedback dialog — shown after a batch finishes processing
# =====================================================================

class FeedbackDialog(tk.Toplevel):
    def __init__(self, master, summary_text=""):
        super().__init__(master)
        self.configure(bg=BG)
        self.title("Processing Complete")
        self.geometry("420x320")
        self.resizable(False, False)

        ttk.Label(self, text="Processing complete!", font=("TkDefaultFont", 12, "bold")).pack(
            anchor="w", padx=14, pady=(16, 2))
        if summary_text:
            ttk.Label(self, text=summary_text, wraplength=390, justify="left").pack(
                anchor="w", padx=14, pady=(0, 10))

        ttk.Label(self, text="How did it go? Any feedback helps improve TigTag.",
                  wraplength=390, justify="left").pack(anchor="w", padx=14, pady=(4, 6))

        self.comment_box = tk.Text(self, height=6, wrap="word")
        style_text(self.comment_box)
        self.comment_box.pack(fill="both", expand=True, padx=14, pady=(0, 10))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=14, pady=(0, 14))
        ttk.Button(btns, text="Skip", command=self.destroy).pack(side="right")
        ttk.Button(btns, text="Submit Feedback", style="Accent.TButton",
                   command=self._submit).pack(side="right", padx=6)

    def _submit(self):
        comment = self.comment_box.get("1.0", "end").strip()
        try:
            with open(FEEDBACK_LOG, "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {comment or '(no comment)'}\n")
        except Exception:
            pass
        messagebox.showinfo(APP_TITLE, "Thanks for your feedback!")
        self.destroy()


# =====================================================================
# Main application window
# =====================================================================

class TaggerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x640")
        apply_theme(self)
        self._set_window_icon()

        if openpyxl is None:
            messagebox.showerror(APP_TITLE, "openpyxl is not installed.\nRun: pip install openpyxl")

        self.config_store = ConfigStore()
        self.file_sheets = {}   # {file_path: {sheet_name: BooleanVar}}
        self.file_order = []

        self._build_menu_row()
        self._build_file_panel()
        self._build_log_panel()

    # ---------- UI construction ----------

    def _set_window_icon(self):
        if not ICON_PATH.exists():
            return
        try:
            self._icon_img = tk.PhotoImage(file=str(ICON_PATH))
            self.iconphoto(True, self._icon_img)
        except Exception:
            pass

    def _build_menu_row(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=10)

        ttk.Button(bar, text="Add Workbook(s)...", style="TopBar.TButton", command=self.add_files).pack(side="left")
        ttk.Button(bar, text="Remove Selected File", style="TopBar.TButton", command=self.remove_selected_file).pack(side="left", padx=6)
        ttk.Separator(bar, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(bar, text="Scan & Configure Keywords", style="TopBar.TButton", command=self.open_detection_panel).pack(side="left")
        ttk.Button(bar, text="Blacklist", style="TopBar.TButton", command=lambda: TagSubtagListDialog(self, self.config_store, "blacklist")).pack(side="left", padx=6)
        ttk.Button(bar, text="Whitelist", style="TopBar.TButton", command=lambda: TagSubtagListDialog(self, self.config_store, "whitelist")).pack(side="left")
        ttk.Button(bar, text="Gender Names", style="TopBar.TButton", command=lambda: GenderNamesDialog(self, self.config_store)).pack(side="left", padx=6)
        ttk.Button(bar, text="File Types", style="TopBar.TButton", command=lambda: ExtensionsDialog(self, self.config_store)).pack(side="left")
        ttk.Separator(bar, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(bar, text="About", style="TopBar.TButton", command=lambda: AboutDialog(self)).pack(side="left")

    def _build_file_panel(self):
        mid = ttk.Frame(self)
        mid.pack(fill="both", expand=True, padx=10)

        left = ttk.LabelFrame(mid, text="Workbooks")
        left.pack(side="left", fill="both", expand=True, padx=(0, 5))
        self.file_listbox = tk.Listbox(left, height=10)
        style_listbox(self.file_listbox)
        self.file_listbox.pack(fill="both", expand=True, padx=5, pady=5)
        self.file_listbox.bind("<<ListboxSelect>>", self._on_file_select)

        right = ttk.LabelFrame(mid, text="Worksheets in selected workbook")
        right.pack(side="left", fill="both", expand=True, padx=(5, 0))
        self.sheet_frame = ScrollableFrame(right, height=180)
        self.sheet_frame.pack(fill="both", expand=True, padx=5, pady=5)

        opts = ttk.Frame(self)
        opts.pack(fill="x", padx=10, pady=(8, 0))
        self.include_folder_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opts, text="Also tag folder-only rows (no file extension) with basic keywords",
                        variable=self.include_folder_var).pack(side="left")

        run_row = ttk.Frame(self)
        run_row.pack(fill="x", padx=10, pady=10)
        self.progress = ttk.Progressbar(run_row, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(run_row, text="Process All", style="Accent.TButton", command=self.process_all).pack(side="right")

    def _build_log_panel(self):
        frame = ttk.LabelFrame(self, text="Log")
        frame.pack(fill="both", expand=False, padx=10, pady=(0, 10))
        self.log_text = tk.Text(frame, height=10, wrap="word", state="disabled")
        style_text(self.log_text)
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)

    # ---------- logging ----------

    def log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    # ---------- file management ----------

    def add_files(self):
        paths = filedialog.askopenfilenames(title="Select Excel workbook(s)",
                                             filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")])
        if not paths:
            return
        for p in paths:
            if p in self.file_sheets:
                continue
            try:
                wb = openpyxl.load_workbook(p, read_only=True)
                sheetnames = wb.sheetnames
                wb.close()
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"Could not open {p}:\n{e}")
                continue
            self.file_sheets[p] = {sn: tk.BooleanVar(value=True) for sn in sheetnames}
            self.file_order.append(p)
            self.file_listbox.insert("end", os.path.basename(p))
        self.log(f"Added {len(paths)} file(s).")

    def remove_selected_file(self):
        sel = self.file_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        path = self.file_order.pop(idx)
        del self.file_sheets[path]
        self.file_listbox.delete(idx)
        for child in self.sheet_frame.inner.winfo_children():
            child.destroy()

    def _on_file_select(self, event):
        for child in self.sheet_frame.inner.winfo_children():
            child.destroy()
        sel = self.file_listbox.curselection()
        if not sel:
            return
        path = self.file_order[sel[0]]
        for sn, var in self.file_sheets[path].items():
            ttk.Checkbutton(self.sheet_frame.inner, text=sn, variable=var).pack(anchor="w", padx=4, pady=1)

    # ---------- keyword detection ----------

    def _gather_all_pairs(self):
        pairs = []
        for path in self.file_order:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            except Exception as e:
                self.log(f"[ERROR] {path}: {e}")
                continue
            selected_sheets = [sn for sn, v in self.file_sheets[path].items() if v.get()]
            for sn in selected_sheets:
                if sn not in wb.sheetnames:
                    continue
                ws = wb[sn]
                for r, a, c in ExcelProcessor.iter_rows(ws):
                    pairs.append((a, c))
            wb.close()
        return pairs

    def open_detection_panel(self):
        if not self.file_order:
            messagebox.showwarning(APP_TITLE, "Add at least one workbook first.")
            return
        self.log("Scanning selected sheets for candidate keywords...")
        pairs = self._gather_all_pairs()
        if not pairs:
            messagebox.showwarning(APP_TITLE, "No data rows found in the selected sheets.")
            return
        candidates = scan_candidates(self.config_store, pairs)
        self.log(f"Found {len(candidates)} candidate keyword(s) across {len(pairs)} row(s).")
        DetectionPanel(self, self.config_store, candidates)

    # ---------- processing ----------

    def process_all(self):
        if not self.file_order:
            messagebox.showwarning(APP_TITLE, "Add at least one workbook first.")
            return
        if not self.config_store.keyword_rules:
            if not messagebox.askyesno(APP_TITLE, "No keywords have been configured yet via "
                                                    "'Scan & Configure Keywords'. Continue anyway "
                                                    "(only 'Audio' + gender + whitelist tags will be written)?"):
                return
        self.progress["value"] = 0
        self.progress["maximum"] = len(self.file_order)
        thread = threading.Thread(target=self._process_worker, daemon=True)
        thread.start()

    def _process_worker(self):
        include_folder = self.include_folder_var.get()
        outputs = []
        for path in self.file_order:
            self.log(f"Processing {os.path.basename(path)} ...")
            selected_sheets = [sn for sn, v in self.file_sheets[path].items() if v.get()]
            try:
                out_path, processed, skipped = process_workbook(
                    path, selected_sheets, self.config_store,
                    include_folder_rows=include_folder, log=self.log,
                )
                self.log(f"[OK] {out_path.name} — {processed} rows tagged, {skipped} skipped.\n")
                outputs.append(str(out_path))
            except Exception as e:
                self.log(f"[ERROR] {path}: {e}\n{traceback.format_exc()}")
            self.progress["value"] += 1
            self.update_idletasks()
        summary = f"Done. {len(outputs)} output file(s) written next to their source files."
        self.log(summary)
        self.after(0, lambda: FeedbackDialog(self, summary_text=summary))


def main():
    app = TaggerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
